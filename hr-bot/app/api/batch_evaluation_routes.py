"""
批量AI评估API路由
- 离线批量调用AI评估
- 自动从共享资源获取JD和简历
- 存储评估结果到缓存目录
"""

import os
import json
import logging
from datetime import datetime
from typing import List, Dict, Optional
from pydantic import BaseModel, Field
from fastapi import APIRouter, HTTPException, BackgroundTasks
import aiohttp

from app.config import get_settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/batch-evaluation", tags=["批量AI评估"])

# 配置
BASE_DIR = "/root/shijingjing/e-employee/hr-bot"
PROJECT_BASE_DIR = os.path.join(BASE_DIR, "data", "interview")

# Qwen3-235B 大模型配置
QWEN_API_URL = "http://180.97.200.118:30071/v1/chat/completions"
QWEN_API_KEY = "z3oK7bN9xPqW2mT8rYvL5tF1cJ4hD6gA0eS2uI3nQk"
QWEN_MODEL = "Qwen/Qwen3-235B-A22B-Instruct-2507"

# 6个评估维度
EVALUATION_DIMENSIONS = [
    "专业能力",
    "工作经验",
    "沟通表达",
    "逻辑思维",
    "学习能力",
    "综合素质"
]


class BatchEvaluationResponse(BaseModel):
    """批量评估响应"""
    success: bool
    message: str
    total: int
    processed: int
    failed: int
    results: List[Dict]


class EvaluationResult(BaseModel):
    """单个候选人评估结果"""
    candidate_name: str
    success: bool
    overall_score: float
    evaluation_level: str
    dimensions: List[dict]
    summary: str
    strengths: List[str]
    weaknesses: List[str]
    recommendations: List[str]
    error: Optional[str] = None


def load_shared_resources(project_name: str) -> Dict:
    """加载项目的共享资源（JD、评价标准）"""
    try:
        resource_path = os.path.join(PROJECT_BASE_DIR, project_name, "_shared_resources.json")
        if os.path.exists(resource_path):
            with open(resource_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    except Exception as e:
        logger.warning(f"[批量评估] 加载共享资源失败: {e}")
        return {}


def load_resume_text(project_name: str, candidate_name: str) -> Optional[str]:
    """加载候选人的简历文本"""
    try:
        project_dir = os.path.join(PROJECT_BASE_DIR, project_name)
        
        # 查找Excel简历文件
        for filename in os.listdir(project_dir):
            if candidate_name in filename and filename.endswith('.xlsx'):
                excel_path = os.path.join(project_dir, filename)
                
                # 读取Excel内容
                import openpyxl
                workbook = openpyxl.load_workbook(excel_path)
                
                resume_text = f"候选人: {candidate_name}\n\n"
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    resume_text += f"【{sheet_name}】\n"
                    
                    for row in sheet.iter_rows(values_only=True):
                        if any(row):
                            row_text = " | ".join([str(cell) for cell in row if cell])
                            if row_text:
                                resume_text += row_text + "\n"
                    
                    resume_text += "\n"
                
                return resume_text
        
        return None
    except Exception as e:
        logger.warning(f"[批量评估] 加载简历失败 {candidate_name}: {e}")
        return None


def load_transcription(project_name: str, candidate_name: str) -> Optional[str]:
    """加载候选人的转录文本"""
    try:
        cache_dir = os.path.join(PROJECT_BASE_DIR, project_name, "transcriptions")
        
        # 首先尝试从 _processing_summary.json 读取
        summary_file = os.path.join(cache_dir, "_processing_summary.json")
        if os.path.exists(summary_file):
            with open(summary_file, 'r', encoding='utf-8') as f:
                summary_data = json.load(f)
            
            for result in summary_data.get("results", []):
                if result.get("candidate_name") == candidate_name:
                    return result.get("transcription", "")
        
        # 尝试从候选人子目录读取
        candidate_dir = os.path.join(cache_dir, candidate_name)
        if os.path.exists(candidate_dir):
            for filename in os.listdir(candidate_dir):
                if filename.endswith('.json'):
                    with open(os.path.join(candidate_dir, filename), 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        return data.get("transcription", "")
        
        return None
    except Exception as e:
        logger.warning(f"[批量评估] 加载转录失败 {candidate_name}: {e}")
        return None


async def call_qwen_model(prompt: str, temperature: float = 0.3) -> str:
    """调用 Qwen3-235B 大模型"""
    try:
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {QWEN_API_KEY}"
        }
        
        payload = {
            "model": QWEN_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": temperature,
            "stream": False
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.post(QWEN_API_URL, headers=headers, json=payload, timeout=300) as response:
                if response.status == 200:
                    result = await response.json()
                    return result.get("choices", [{}])[0].get("message", {}).get("content", "")
                else:
                    error_text = await response.text()
                    logger.error(f"[批量评估] Qwen API错误: {response.status}, {error_text}")
                    return ""
                    
    except Exception as e:
        logger.error(f"[批量评估] 调用Qwen模型失败: {e}")
        return ""


async def evaluate_candidate_with_ai(
    candidate_name: str,
    jd_content: str,
    resume_content: str,
    transcript: str
) -> Optional[Dict]:
    """使用AI评估候选人"""
    
    prompt = f"""你是一位专业HR面试官，请根据以下信息对候选人进行全面的面试评价。

## 岗位JD
{jd_content[:2000]}

## 候选人简历
{resume_content[:2000]}

## 面试录音转录
{transcript[:3000]}

请根据以上信息，对候选人进行综合评价：

## 评分维度说明

请根据候选人在面试中的表现，按以下6个维度进行评分（0-100分）：

1. **专业能力**：考察候选人的专业技能和知识储备
2. **工作经验**：考察候选人的工作经历与岗位的匹配度
3. **沟通表达**：考察候选人的语言表达和沟通能力
4. **逻辑思维**：考察候选人的思维逻辑和分析能力
5. **学习能力**：考察候选人的学习潜力和成长空间
6. **综合素质**：考察候选人的团队协作、职业素养等

请按以下JSON格式输出结果：

```json
{{
    "overall_score": 85,
    "evaluation_level": "优秀/良好/一般/较差",
    "dimensions": [
        {{"name": "专业能力", "score": 85, "weight": 18, "analysis": "详细分析说明"}},
        {{"name": "工作经验", "score": 80, "weight": 18, "analysis": "详细分析说明"}},
        {{"name": "沟通表达", "score": 82, "weight": 16, "analysis": "详细分析说明"}},
        {{"name": "逻辑思维", "score": 78, "weight": 16, "analysis": "详细分析说明"}},
        {{"name": "学习能力", "score": 85, "weight": 16, "analysis": "详细分析说明"}},
        {{"name": "综合素质", "score": 80, "weight": 16, "analysis": "详细分析说明"}}
    ],
    "summary": "综合评价总结（100-200字）",
    "strengths": ["优势1", "优势2", "优势3"],
    "weaknesses": ["不足1", "不足2"],
    "recommendations": ["建议1", "建议2"]
}}
```

注意：
1. overall_score为0-100的综合评分
2. evaluation_level根据分数划分：90-100优秀，80-89良好，60-79一般，60以下较差
3. 各维度权重总和必须为100%
4. 只输出JSON格式内容，不要有其他说明文字"""

    content = await call_qwen_model(prompt, temperature=0.3)
    
    try:
        # 解析JSON
        if "```json" in content:
            json_str = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            json_str = content.split("```")[1].strip()
        else:
            json_str = content.strip()
        
        result = json.loads(json_str)
        
        # 添加候选人姓名
        result["candidate_name"] = candidate_name
        
        logger.info(f"[批量评估] {candidate_name} 评估完成，综合得分: {result.get('overall_score')}")
        return result
        
    except json.JSONDecodeError as e:
        logger.error(f"[批量评估] 评估结果解析失败 {candidate_name}: {e}")
        return None


def save_evaluation_cache(project_name: str, candidate_name: str, evaluation: Dict):
    """保存评估结果到缓存文件"""
    try:
        evaluations_dir = os.path.join(PROJECT_BASE_DIR, project_name, "evaluations")
        os.makedirs(evaluations_dir, exist_ok=True)
        
        cache_file = os.path.join(evaluations_dir, f"{candidate_name}_evaluation.json")
        
        cache_data = {
            "candidate_name": candidate_name,
            "project": project_name,
            "evaluation": evaluation,
            "cached_at": datetime.now().isoformat()
        }
        
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"[批量评估] 已保存缓存: {cache_file}")
        return True
        
    except Exception as e:
        logger.error(f"[批量评估] 保存缓存失败 {candidate_name}: {e}")
        return False


@router.post("/run", response_model=BatchEvaluationResponse)
async def run_batch_evaluation(
    project_name: str,
    skip_existing: bool = True,
    max_candidates: Optional[int] = None
):
    """
    运行批量AI评估
    
    - 自动获取项目所有候选人
    - 从共享资源获取JD
    - 从Excel文件获取简历
    - 从转录缓存获取面试文本
    - 调用AI进行评估
    - 保存结果到缓存目录
    
    Args:
        project_name: 项目名称
        skip_existing: 是否跳过已有缓存的候选人
        max_candidates: 最大评估数量（用于测试）
    """
    logger.info(f"[批量评估] 开始批量评估: {project_name}")
    
    try:
        # 1. 加载共享资源（JD）
        shared_resources = load_shared_resources(project_name)
        jd_content = shared_resources.get("job_description", "")
        
        if not jd_content:
            return BatchEvaluationResponse(
                success=False,
                message="未找到JD内容，请先上传JD文件",
                total=0,
                processed=0,
                failed=0,
                results=[]
            )
        
        # 2. 获取所有候选人
        from app.api.interview_batch_routes import get_candidates
        candidates_data = await get_candidates(project_name)
        candidates = candidates_data.get("candidates", [])
        
        if not candidates:
            return BatchEvaluationResponse(
                success=False,
                message="未找到候选人",
                total=0,
                processed=0,
                failed=0,
                results=[]
            )
        
        # 3. 筛选需要评估的候选人
        candidates_to_evaluate = []
        for candidate in candidates:
            name = candidate.get("name", "")
            has_transcription = candidate.get("has_transcription", False)
            has_evaluation = candidate.get("has_evaluation", False)
            
            if not has_transcription:
                logger.info(f"[批量评估] 跳过 {name}: 无转录数据")
                continue
            
            if skip_existing and has_evaluation:
                logger.info(f"[批量评估] 跳过 {name}: 已有评估缓存")
                continue
            
            candidates_to_evaluate.append(candidate)
        
        # 限制评估数量
        if max_candidates and len(candidates_to_evaluate) > max_candidates:
            candidates_to_evaluate = candidates_to_evaluate[:max_candidates]
        
        logger.info(f"[批量评估] 需要评估 {len(candidates_to_evaluate)} 个候选人")
        
        # 4. 批量评估
        results = []
        success_count = 0
        failed_count = 0
        
        for i, candidate in enumerate(candidates_to_evaluate, 1):
            name = candidate.get("name", "")
            logger.info(f"[批量评估] [{i}/{len(candidates_to_evaluate)}] 评估: {name}")
            
            try:
                # 加载简历和转录
                resume_content = load_resume_text(project_name, name)
                transcript = load_transcription(project_name, name)
                
                if not resume_content:
                    logger.warning(f"[批量评估] {name}: 未找到简历")
                    resume_content = "未找到简历文件"
                
                if not transcript:
                    logger.warning(f"[批量评估] {name}: 未找到转录")
                    results.append({
                        "candidate_name": name,
                        "success": False,
                        "error": "未找到转录数据"
                    })
                    failed_count += 1
                    continue
                
                # 调用AI评估
                evaluation = await evaluate_candidate_with_ai(
                    candidate_name=name,
                    jd_content=jd_content,
                    resume_content=resume_content,
                    transcript=transcript
                )
                
                if evaluation:
                    # 保存缓存
                    save_evaluation_cache(project_name, name, evaluation)
                    
                    results.append({
                        "candidate_name": name,
                        "success": True,
                        "overall_score": evaluation.get("overall_score"),
                        "evaluation_level": evaluation.get("evaluation_level"),
                        "dimensions": evaluation.get("dimensions", [])
                    })
                    success_count += 1
                else:
                    results.append({
                        "candidate_name": name,
                        "success": False,
                        "error": "AI评估失败"
                    })
                    failed_count += 1
                
                # 添加延迟避免API过载
                if i < len(candidates_to_evaluate):
                    await asyncio.sleep(1)
                    
            except Exception as e:
                logger.error(f"[批量评估] 评估失败 {name}: {e}")
                results.append({
                    "candidate_name": name,
                    "success": False,
                    "error": str(e)
                })
                failed_count += 1
        
        # 5. 返回结果
        return BatchEvaluationResponse(
            success=True,
            message=f"批量评估完成: 成功{success_count}, 失败{failed_count}",
            total=len(candidates_to_evaluate),
            processed=success_count,
            failed=failed_count,
            results=results
        )
        
    except Exception as e:
        logger.error(f"[批量评估] 批量评估失败: {e}")
        return BatchEvaluationResponse(
            success=False,
            message=f"批量评估失败: {str(e)}",
            total=0,
            processed=0,
            failed=0,
            results=[]
        )


@router.get("/status/{project_name}")
async def get_evaluation_status(project_name: str):
    """获取项目的评估状态"""
    try:
        from app.api.interview_batch_routes import get_candidates
        candidates_data = await get_candidates(project_name)
        candidates = candidates_data.get("candidates", [])
        
        total = len(candidates)
        evaluated = sum(1 for c in candidates if c.get("has_evaluation", False))
        with_transcription = sum(1 for c in candidates if c.get("has_transcription", False))
        
        return {
            "success": True,
            "project": project_name,
            "total_candidates": total,
            "evaluated": evaluated,
            "pending": total - evaluated,
            "with_transcription": with_transcription,
            "progress": f"{evaluated}/{total} ({evaluated/total*100:.1f}%)" if total > 0 else "0%"
        }
        
    except Exception as e:
        return {
            "success": False,
            "message": f"获取状态失败: {str(e)}"
        }
