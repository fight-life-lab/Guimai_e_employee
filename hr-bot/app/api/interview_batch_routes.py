"""
批量面试处理API路由
支持批量转录、读取Excel简历、共享问答对和招聘JD
"""

import os
import json
import asyncio
import aiohttp
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Any
from pydantic import BaseModel, Field
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, BackgroundTasks
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/interview-batch", tags=["批量面试处理"])

# ============ 配置 ============
WHISPER_API_URL = "http://localhost:8003/transcribe"
PROJECT_BASE_DIR = Path("/root/shijingjing/e-employee/hr-bot/data/interview")

# ============ 数据模型 ============

class BatchTranscribeRequest(BaseModel):
    """批量转录请求"""
    project_name: str = Field(..., description="项目名称，如：20260401战略招聘")
    candidate_names: Optional[List[str]] = Field(None, description="指定候选人列表，为空则处理所有")


class BatchTranscribeResponse(BaseModel):
    """批量转录响应"""
    success: bool
    total: int
    processed: int
    cached: int
    failed: int
    results: List[Dict]


class CandidateResume(BaseModel):
    """候选人简历信息"""
    name: str
    file_name: str
    basic_info: Dict[str, Any]
    education: List[Dict]
    work_experience: List[Dict]
    projects: List[Dict]
    skills: List[str]


class SharedResources(BaseModel):
    """共享资源"""
    job_description: str = Field("", description="招聘JD")
    interview_questions: List[Dict] = Field([], description="结构化面试问题")
    evaluation_criteria: str = Field("", description="评价标准")


class BatchEvaluationRequest(BaseModel):
    """批量评估请求"""
    project_name: str
    candidate_names: Optional[List[str]] = None
    use_shared_jd: bool = True
    use_shared_questions: bool = True


# ============ 工具函数 ============

def get_project_dir(project_name: str) -> Path:
    """获取项目目录"""
    return PROJECT_BASE_DIR / project_name


def get_transcription_cache_dir(project_name: str) -> Path:
    """获取转录缓存目录"""
    cache_dir = get_project_dir(project_name) / "transcriptions"
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def get_shared_resources_path(project_name: str) -> Path:
    """获取共享资源文件路径"""
    return get_project_dir(project_name) / "_shared_resources.json"


def load_shared_resources(project_name: str) -> SharedResources:
    """加载共享资源"""
    resource_path = get_shared_resources_path(project_name)
    if resource_path.exists():
        with open(resource_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return SharedResources(**data)
    return SharedResources()


def save_shared_resources(project_name: str, resources: SharedResources):
    """保存共享资源"""
    resource_path = get_shared_resources_path(project_name)
    with open(resource_path, 'w', encoding='utf-8') as f:
        json.dump(resources.dict(), f, ensure_ascii=False, indent=2)


def extract_candidate_name_from_filename(filename: str) -> str:
    """从文件名提取候选人姓名"""
    # 格式: 04月10日_10黄俊华.aac 或 新国脉数字文化股份有限公司招聘报名表（仇天硕）.xlsx
    try:
        # 尝试匹配录音文件名格式
        if '_' in filename:
            name_part = filename.split('_')[-1]
            name = name_part.split('.')[0]
            # 移除数字前缀
            for i, char in enumerate(name):
                if not char.isdigit():
                    return name[i:]
        
        # 尝试匹配报名表文件名格式
        if '（' in filename and '）' in filename:
            start = filename.find('（') + 1
            end = filename.find('）')
            return filename[start:end]
        
        return filename.split('.')[0]
    except:
        return filename


def parse_excel_resume(excel_path: Path) -> CandidateResume:
    """解析Excel格式的简历"""
    try:
        workbook = openpyxl.load_workbook(excel_path)
        
        # 提取候选人姓名
        candidate_name = extract_candidate_name_from_filename(excel_path.name)
        
        # 解析基本信息
        basic_info = {}
        education = []
        work_experience = []
        projects = []
        skills = []
        
        # 遍历所有sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 根据sheet名称解析不同部分
            if '基本信息' in sheet_name or 'basic' in sheet_name.lower():
                basic_info = parse_basic_info(sheet)
            elif '教育' in sheet_name or 'education' in sheet_name.lower():
                education = parse_education(sheet)
            elif '工作' in sheet_name or 'work' in sheet_name.lower():
                work_experience = parse_work_experience(sheet)
            elif '项目' in sheet_name or 'project' in sheet_name.lower():
                projects = parse_projects(sheet)
            elif '技能' in sheet_name or 'skill' in sheet_name.lower():
                skills = parse_skills(sheet)
        
        # 如果没有特定sheet，尝试从第一个sheet解析
        if not basic_info and workbook.sheetnames:
            basic_info = parse_basic_info(workbook.active)
        
        return CandidateResume(
            name=candidate_name,
            file_name=excel_path.name,
            basic_info=basic_info,
            education=education,
            work_experience=work_experience,
            projects=projects,
            skills=skills
        )
    except Exception as e:
        logger.error(f"解析Excel简历失败 {excel_path}: {e}")
        return CandidateResume(
            name=extract_candidate_name_from_filename(excel_path.name),
            file_name=excel_path.name,
            basic_info={},
            education=[],
            work_experience=[],
            projects=[],
            skills=[]
        )


def parse_basic_info(sheet) -> Dict:
    """解析基本信息"""
    info = {}
    for row in sheet.iter_rows(values_only=True):
        if len(row) >= 2 and row[0] and row[1]:
            key = str(row[0]).strip()
            value = str(row[1]).strip()
            if key and value:
                info[key] = value
    return info


def parse_education(sheet) -> List[Dict]:
    """解析教育经历"""
    education = []
    headers = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c else "" for c in row]
        else:
            if any(row):
                edu = {}
                for j, value in enumerate(row):
                    if j < len(headers) and value:
                        edu[headers[j]] = str(value)
                if edu:
                    education.append(edu)
    return education


def parse_work_experience(sheet) -> List[Dict]:
    """解析工作经历"""
    return parse_education(sheet)  # 结构相同


def parse_projects(sheet) -> List[Dict]:
    """解析项目经历"""
    return parse_education(sheet)  # 结构相同


def parse_skills(sheet) -> List[str]:
    """解析技能"""
    skills = []
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell:
                skills.append(str(cell))
    return skills


async def transcribe_single_audio(audio_path: Path, cache_dir: Path) -> Dict:
    """转录单个音频文件"""
    file_name = audio_path.name
    candidate_name = extract_candidate_name_from_filename(file_name)
    cache_file = cache_dir / f"{file_name}.json"
    
    # 检查缓存
    if cache_file.exists():
        with open(cache_file, 'r', encoding='utf-8') as f:
            cached = json.load(f)
            cached['cached'] = True
            return cached
    
    # 调用Whisper API
    try:
        async with aiohttp.ClientSession() as session:
            data = aiohttp.FormData()
            data.add_field('audio_file',
                         open(audio_path, 'rb'),
                         filename=file_name,
                         content_type='audio/aac')
            
            async with session.post(WHISPER_API_URL, data=data, timeout=aiohttp.ClientTimeout(total=300)) as response:
                if response.status == 200:
                    result = await response.json()
                    
                    transcription_data = {
                        "file_name": file_name,
                        "candidate_name": candidate_name,
                        "transcription": result.get("text", ""),
                        "segments": result.get("segments", []),
                        "language": result.get("language", "zh"),
                        "processed_at": datetime.now().isoformat(),
                        "cached": False
                    }
                    
                    # 保存缓存
                    with open(cache_file, 'w', encoding='utf-8') as f:
                        json.dump(transcription_data, f, ensure_ascii=False, indent=2)
                    
                    return transcription_data
                else:
                    error_text = await response.text()
                    return {
                        "file_name": file_name,
                        "candidate_name": candidate_name,
                        "error": f"API错误: {response.status}",
                        "processed_at": datetime.now().isoformat()
                    }
    except Exception as e:
        return {
            "file_name": file_name,
            "candidate_name": candidate_name,
            "error": str(e),
            "processed_at": datetime.now().isoformat()
        }


# ============ API路由 ============

@router.post("/transcribe-batch", response_model=BatchTranscribeResponse)
async def batch_transcribe(request: BatchTranscribeRequest, background_tasks: BackgroundTasks):
    """
    批量转录面试录音
    
    - 自动识别项目目录下的所有音频文件
    - 支持缓存，避免重复转录
    - 异步处理，返回处理状态
    """
    project_dir = get_project_dir(request.project_name)
    cache_dir = get_transcription_cache_dir(request.project_name)
    
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail=f"项目不存在: {request.project_name}")
    
    # 获取所有音频文件
    audio_extensions = {'.aac', '.mp3', '.wav', '.m4a', '.mp4'}
    audio_files = [f for f in project_dir.iterdir() 
                   if f.is_file() and f.suffix.lower() in audio_extensions]
    
    # 如果指定了候选人，只处理匹配的
    if request.candidate_names:
        audio_files = [f for f in audio_files 
                      if any(name in f.name for name in request.candidate_names)]
    
    results = []
    cached_count = 0
    failed_count = 0
    
    for audio_file in audio_files:
        result = await transcribe_single_audio(audio_file, cache_dir)
        results.append(result)
        
        if result.get('cached'):
            cached_count += 1
        elif result.get('error'):
            failed_count += 1
    
    return BatchTranscribeResponse(
        success=True,
        total=len(audio_files),
        processed=len(audio_files) - cached_count - failed_count,
        cached=cached_count,
        failed=failed_count,
        results=results
    )


@router.get("/resumes/{project_name}")
async def get_resumes(project_name: str):
    """
    获取项目中所有候选人的简历信息
    
    - 自动解析所有Excel格式的报名表
    - 返回结构化的简历数据
    """
    project_dir = get_project_dir(project_name)
    
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail=f"项目不存在: {project_name}")
    
    # 查找所有Excel文件
    excel_files = list(project_dir.glob("*.xlsx")) + list(project_dir.glob("*.xls"))
    
    resumes = []
    for excel_file in excel_files:
        resume = parse_excel_resume(excel_file)
        resumes.append(resume.dict())
    
    return {
        "success": True,
        "total": len(resumes),
        "resumes": resumes
    }


@router.get("/shared-resources/{project_name}")
async def get_shared_resources_api(project_name: str):
    """获取项目的共享资源（JD、问答对）"""
    resources = load_shared_resources(project_name)
    return {
        "success": True,
        "project": project_name,
        "resources": resources.dict()
    }


@router.post("/shared-resources/{project_name}")
async def update_shared_resources(
    project_name: str,
    job_description: Optional[str] = Form(None),
    evaluation_criteria: Optional[str] = Form(None),
    questions_file: Optional[UploadFile] = File(None)
):
    """
    更新项目的共享资源
    
    - 上传招聘JD
    - 上传结构化面试问题Excel
    - 设置评价标准
    """
    resources = load_shared_resources(project_name)
    
    # 更新JD
    if job_description:
        resources.job_description = job_description
    
    # 更新评价标准
    if evaluation_criteria:
        resources.evaluation_criteria = evaluation_criteria
    
    # 解析面试问题Excel
    if questions_file:
        try:
            # 保存上传的文件
            temp_path = get_project_dir(project_name) / "_temp_questions.xlsx"
            with open(temp_path, 'wb') as f:
                content = await questions_file.read()
                f.write(content)
            
            # 解析Excel
            workbook = openpyxl.load_workbook(temp_path)
            sheet = workbook.active
            
            questions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[0] and row[1]:
                    questions.append({
                        "category": str(row[0]) if row[0] else "",
                        "question": str(row[1]) if row[1] else "",
                        "evaluation_points": str(row[2]) if len(row) > 2 and row[2] else ""
                    })
            
            resources.interview_questions = questions
            
            # 删除临时文件
            temp_path.unlink()
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"解析面试问题文件失败: {e}")
    
    # 保存资源
    save_shared_resources(project_name, resources)
    
    return {
        "success": True,
        "message": "共享资源已更新",
        "resources": resources.dict()
    }


@router.get("/transcriptions/{project_name}")
async def get_transcriptions(project_name: str, candidate_name: Optional[str] = None):
    """获取项目的转录结果
    
    支持三种缓存结构：
    1. 旧结构: transcriptions/{文件名}.json
    2. 新结构: transcriptions/{候选人姓名}/{文件名}.json
    3. 汇总结构: transcriptions/_processing_summary.json
    """
    cache_dir = get_transcription_cache_dir(project_name)
    
    if not cache_dir.exists():
        return {"success": True, "transcriptions": []}
    
    transcriptions = []
    
    # 首先尝试从 _processing_summary.json 读取转录数据
    summary_file = cache_dir / "_processing_summary.json"
    if summary_file.exists():
        try:
            with open(summary_file, 'r', encoding='utf-8') as f:
                summary_data = json.load(f)
            
            results = summary_data.get("results", [])
            for result in results:
                name = result.get("candidate_name", "")
                # 跳过没有候选人姓名的条目
                if not name:
                    continue
                # 如果指定了候选人，只返回匹配的
                if candidate_name and name != candidate_name:
                    continue
                
                transcriptions.append({
                    "candidate_name": name,
                    "transcription": result.get("transcription", ""),
                    "audio_file": result.get("audio_file", ""),
                    "cached_at": result.get("cached_at", "")
                })
            
            if transcriptions:
                return {
                    "success": True,
                    "total": len(transcriptions),
                    "transcriptions": transcriptions
                }
        except Exception as e:
            logger.warning(f"读取汇总文件失败 {summary_file}: {e}")
    
    # 如果没有汇总文件或读取失败，则递归查找所有 .json 文件
    def find_cache_files(directory: Path):
        """递归查找所有缓存文件"""
        cache_files = []
        for item in directory.iterdir():
            if item.is_dir():
                # 递归查找子目录
                cache_files.extend(find_cache_files(item))
            elif item.is_file() and item.suffix == '.json' and not item.name.startswith('_'):
                cache_files.append(item)
        return cache_files
    
    cache_files = find_cache_files(cache_dir)
    
    for cache_file in cache_files:
        try:
            with open(cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                # 跳过没有候选人姓名的数据
                if not data.get('candidate_name'):
                    continue
                
                # 如果指定了候选人，只返回匹配的
                if candidate_name and data.get('candidate_name') != candidate_name:
                    continue
                
                transcriptions.append(data)
        except Exception as e:
            logger.warning(f"读取缓存文件失败 {cache_file}: {e}")
    
    return {
        "success": True,
        "total": len(transcriptions),
        "transcriptions": transcriptions
    }


def load_evaluation_cache_for_candidate(project_name: str, candidate_name: str) -> Optional[Dict]:
    """加载候选人的AI评估缓存"""
    try:
        evaluations_dir = get_project_dir(project_name) / "evaluations"
        if not evaluations_dir.exists():
            return None
        
        cache_file = evaluations_dir / f"{candidate_name}_evaluation.json"
        if not cache_file.exists():
            return None
        
        with open(cache_file, 'r', encoding='utf-8') as f:
            cache_data = json.load(f)
        
        # 兼容两种格式：
        # 1. 旧格式: { "evaluation": { ... } }
        # 2. 新格式: { "candidate_performance": { ... }, "salary_match": { ... } }
        if "evaluation" in cache_data:
            return cache_data.get("evaluation")
        elif "candidate_performance" in cache_data:
            # 新格式，直接返回整个数据（去掉metadata）
            return {
                "candidate_performance": cache_data.get("candidate_performance"),
                "salary_match": cache_data.get("salary_match"),
                "jd_requirements": cache_data.get("jd_requirements"),
                "qa_summary": cache_data.get("qa_summary")
            }
        else:
            # 如果没有标准字段，返回整个数据
            return cache_data
    except Exception as e:
        logger.warning(f"加载评估缓存失败 {candidate_name}: {e}")
        return None


@router.get("/candidates/{project_name}")
async def get_candidates(project_name: str):
    """
    获取项目所有候选人综合信息
    
    - 合并简历、转录、AI评估信息
    - 返回完整的候选人列表
    """
    project_dir = get_project_dir(project_name)
    
    if not project_dir.exists():
        raise HTTPException(status_code=404, detail=f"项目不存在: {project_name}")
    
    # 获取所有候选人姓名（从音频文件和Excel文件）
    candidate_names = set()
    
    # 从音频文件提取
    for f in project_dir.iterdir():
        if f.is_file() and f.suffix.lower() in {'.aac', '.mp3', '.wav', '.m4a'}:
            name = extract_candidate_name_from_filename(f.name)
            candidate_names.add(name)
    
    # 从Excel文件提取
    for f in project_dir.glob("*.xlsx"):
        name = extract_candidate_name_from_filename(f.name)
        candidate_names.add(name)
    
    # 获取简历信息
    resumes_response = await get_resumes(project_name)
    resumes_map = {r['name']: r for r in resumes_response['resumes']}
    
    # 获取转录信息
    transcriptions_response = await get_transcriptions(project_name)
    transcriptions_map = {t['candidate_name']: t for t in transcriptions_response['transcriptions'] if t.get('candidate_name')}
    
    # 合并信息
    candidates = []
    for name in candidate_names:
        # 加载AI评估缓存
        evaluation = load_evaluation_cache_for_candidate(project_name, name)
        
        candidate = {
            "name": name,
            "has_resume": name in resumes_map,
            "has_transcription": name in transcriptions_map,
            "has_evaluation": evaluation is not None,
            "resume": resumes_map.get(name),
            "transcription": transcriptions_map.get(name),
            "evaluation": evaluation
        }
        candidates.append(candidate)
    
    return {
        "success": True,
        "total": len(candidates),
        "candidates": candidates
    }


@router.delete("/candidates/{project_name}/{candidate_name}")
async def delete_candidate(project_name: str, candidate_name: str):
    """
    删除候选人及其所有相关数据
    
    - 删除转录缓存
    - 删除评估缓存
    - 删除简历文件（可选）
    """
    try:
        project_dir = get_project_dir(project_name)
        
        if not project_dir.exists():
            raise HTTPException(status_code=404, detail=f"项目不存在: {project_name}")
        
        deleted_items = []
        
        # 1. 删除转录缓存
        transcriptions_dir = project_dir / "transcriptions" / candidate_name
        if transcriptions_dir.exists():
            import shutil
            shutil.rmtree(transcriptions_dir)
            deleted_items.append(f"转录缓存: {transcriptions_dir}")
        
        # 2. 删除评估缓存
        evaluations_dir = project_dir / "evaluations"
        if evaluations_dir.exists():
            for eval_file in evaluations_dir.glob(f"{candidate_name}_evaluation.json"):
                eval_file.unlink()
                deleted_items.append(f"评估缓存: {eval_file}")
        
        # 3. 删除音频文件
        for audio_ext in ['.aac', '.mp3', '.wav', '.m4a']:
            for audio_file in project_dir.glob(f"*{candidate_name}*{audio_ext}"):
                audio_file.unlink()
                deleted_items.append(f"音频文件: {audio_file}")
        
        # 4. 删除Excel简历文件
        for excel_file in project_dir.glob(f"*{candidate_name}*.xlsx"):
            excel_file.unlink()
            deleted_items.append(f"简历文件: {excel_file}")
        
        logger.info(f"[批量面试] 删除候选人 {candidate_name}: {deleted_items}")
        
        return {
            "success": True,
            "message": f"已删除候选人: {candidate_name}",
            "deleted_items": deleted_items
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[批量面试] 删除候选人失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


# ============ QA提取接口 ============

class QAExtractRequest(BaseModel):
    """QA提取请求"""
    candidate_name: str = Field(..., description="候选人姓名")
    project: str = Field(..., description="项目名称")
    transcription: str = Field(..., description="转录文本")
    questions: Optional[str] = Field(None, description="面试问题列表（JSON格式）")


class QAExtractResponse(BaseModel):
    """QA提取响应"""
    success: bool
    qa_pairs: List[Dict]
    message: str


@router.post("/qa-extract", response_model=QAExtractResponse)
async def extract_qa_from_transcription(
    candidate_name: str = Form(..., description="候选人姓名"),
    project: str = Form(..., description="项目名称"),
    transcription: str = Form(..., description="转录文本"),
    questions: Optional[str] = Form(None, description="面试问题列表（JSON格式）")
):
    """
    从转录文本中提取问题-回答对
    
    支持两种方式：
    1. 提供面试问题列表，从转录文本中匹配回答
    2. 不提供问题列表，自动识别转录文本中的问答对
    """
    try:
        logger.info(f"[QA提取] 开始提取 {candidate_name} 的QA数据")
        
        if not transcription or len(transcription.strip()) < 10:
            return QAExtractResponse(
                success=False,
                qa_pairs=[],
                message="转录文本为空或太短"
            )
        
        qa_pairs = []
        
        # 解析面试问题
        interview_questions = []
        if questions:
            try:
                interview_questions = json.loads(questions)
                if not isinstance(interview_questions, list):
                    interview_questions = []
            except:
                interview_questions = []
        
        if interview_questions:
            # 方式1：根据提供的问题列表匹配回答
            logger.info(f"[QA提取] 使用提供的 {len(interview_questions)} 个问题进行匹配")
            
            # 简单匹配：将转录文本分段，尝试匹配问题
            transcript_segments = split_transcription(transcription)
            
            for i, q in enumerate(interview_questions[:8]):  # 最多处理8个问题
                question_text = q.get('question', '') if isinstance(q, dict) else str(q)
                category = q.get('category', '通用') if isinstance(q, dict) else '通用'
                evaluation_points = q.get('evaluation_points', '') if isinstance(q, dict) else ''
                
                if question_text:
                    # 尝试在转录文本中找到相关回答
                    answer = extract_answer_for_question(transcription, question_text, i, len(interview_questions))
                    
                    qa_pairs.append({
                        "question": question_text,
                        "answer": answer,
                        "category": category,
                        "evaluation_points": evaluation_points,
                        "start_time": 0,
                        "end_time": 0
                    })
        else:
            # 方式2：自动识别问答对
            logger.info("[QA提取] 自动识别转录文本中的问答对")
            qa_pairs = auto_extract_qa_pairs(transcription)
        
        # 保存QA结果到文件
        await save_qa_result(project, candidate_name, qa_pairs, interview_questions)
        
        logger.info(f"[QA提取] {candidate_name} 提取完成，共 {len(qa_pairs)} 个QA对")
        
        return QAExtractResponse(
            success=True,
            qa_pairs=qa_pairs,
            message=f"成功提取 {len(qa_pairs)} 个问题-回答对"
        )
        
    except Exception as e:
        logger.error(f"[QA提取] 失败: {e}")
        return QAExtractResponse(
            success=False,
            qa_pairs=[],
            message=f"提取失败: {str(e)}"
        )


def split_transcription(transcription: str) -> List[str]:
    """将转录文本分割成段落"""
    # 按常见分隔符分割
    segments = []
    
    # 先按句号、问号、感叹号分割
    import re
    sentences = re.split(r'[。！？\n]+', transcription)
    
    # 合并短句成长段落
    current_segment = ""
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 5:
            continue
        
        if len(current_segment) < 200:
            current_segment += sent + "。"
        else:
            if current_segment:
                segments.append(current_segment)
            current_segment = sent + "。"
    
    if current_segment:
        segments.append(current_segment)
    
    return segments if segments else [transcription]


def extract_answer_for_question(transcription: str, question: str, question_index: int, total_questions: int) -> str:
    """从转录文本中提取特定问题的回答"""
    
    # 简单策略：根据问题位置将转录文本分段
    # 每个问题分配大致相等的文本长度
    
    if total_questions <= 0:
        total_questions = 1
    
    # 计算每个问题应该分配的文本长度
    avg_length = len(transcription) // total_questions
    start_pos = question_index * avg_length
    end_pos = start_pos + avg_length + 500  # 多给一些重叠
    
    if end_pos > len(transcription):
        end_pos = len(transcription)
    
    answer = transcription[start_pos:end_pos].strip()
    
    # 清理回答文本
    # 移除开头可能的重复问题词
    question_keywords = ['问题', '请', '介绍一下', '描述', '说说']
    for keyword in question_keywords:
        if answer.startswith(keyword):
            answer = answer[len(keyword):].strip()
    
    # 限制回答长度
    if len(answer) > 1500:
        answer = answer[:1500] + "..."
    
    return answer if answer else "（未找到明确回答）"


def auto_extract_qa_pairs(transcription: str) -> List[Dict]:
    """自动从转录文本中提取问答对"""
    qa_pairs = []
    
    # 简单启发式：假设文本中包含问答模式
    # 查找可能的问题（包含疑问词或问号）
    import re
    
    # 常见疑问词
    question_patterns = [
        r'(请|能|可以).*?(介绍|描述|说说|谈谈|分享).*?(吗|？|\?)',
        r'(什么|怎么|如何|为什么|哪里|谁|何时).*?(？|\?)',
        r'(您|你).*?(经验|看法|想法|观点|建议).*?(？|\?)',
    ]
    
    # 将文本分段
    segments = split_transcription(transcription)
    
    for i, segment in enumerate(segments[:5]):  # 最多取5段
        # 尝试识别问题
        is_question = False
        for pattern in question_patterns:
            if re.search(pattern, segment):
                is_question = True
                break
        
        if is_question or i == 0:  # 第一段通常包含自我介绍
            # 提取问题（简化处理，用前30字作为问题描述）
            question_text = segment[:30] + "..." if len(segment) > 30 else segment
            
            # 回答就是当前段落
            answer = segment
            
            qa_pairs.append({
                "question": f"问题{i+1}: {question_text}",
                "answer": answer,
                "category": "自动识别",
                "evaluation_points": "",
                "start_time": 0,
                "end_time": 0
            })
    
    # 如果没有识别到任何问题，创建一个通用QA
    if not qa_pairs:
        qa_pairs.append({
            "question": "面试交流",
            "answer": transcription[:1000],
            "category": "通用",
            "evaluation_points": "综合考察候选人的表达能力和经验",
            "start_time": 0,
            "end_time": 0
        })
    
    return qa_pairs


async def save_qa_result(project: str, candidate_name: str, qa_pairs: List[Dict], questions: List):
    """保存QA结果到文件"""
    try:
        project_dir = get_project_dir(project)
        qa_dir = project_dir / "qa"
        qa_dir.mkdir(parents=True, exist_ok=True)
        
        qa_file = qa_dir / f"{candidate_name}.json"
        
        qa_data = {
            "candidate_name": candidate_name,
            "qa_pairs": qa_pairs,
            "questions": questions,
            "processed_at": datetime.now().isoformat()
        }
        
        with open(qa_file, 'w', encoding='utf-8') as f:
            json.dump(qa_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"[QA提取] 已保存QA结果到: {qa_file}")
        
    except Exception as e:
        logger.error(f"[QA提取] 保存QA结果失败: {e}")
