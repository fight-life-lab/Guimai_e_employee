"""
共享工具模块 - 员工和干部面试评估通用工具函数

包含：
1. 文件操作工具
2. 缓存管理工具
3. 问答对提取工具
4. 配置管理工具
"""

import os
import json
import re
import logging
from datetime import datetime
from typing import List, Optional, Dict, Any

logger = logging.getLogger(__name__)


# ============ 目录管理 ============

def get_base_interview_dir() -> str:
    """获取面试数据基础目录"""
    # 优先使用本地开发路径
    local_path = "/Users/shijingjing/Desktop/GuomaiProject/e-employee/hr-bot/data/interview"
    if os.path.exists(local_path):
        return local_path
    # 远程服务器路径
    return "/root/shijingjing/e-employee/hr-bot/data/interview"


def get_type_dir(evaluation_type: str) -> str:
    """获取评估类型目录（员工/干部）"""
    type_dir = os.path.join(get_base_interview_dir(), "员工" if evaluation_type == "employee" else "干部")
    os.makedirs(type_dir, exist_ok=True)
    return type_dir


def get_project_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取项目目录"""
    if evaluation_type:
        type_dir = get_type_dir(evaluation_type)
        project_dir = os.path.join(type_dir, project_name)
    else:
        employee_path = os.path.join(get_base_interview_dir(), "员工", project_name)
        cadre_path = os.path.join(get_base_interview_dir(), "干部", project_name)
        if os.path.exists(employee_path):
            project_dir = employee_path
        elif os.path.exists(cadre_path):
            project_dir = cadre_path
        else:
            project_dir = os.path.join(get_base_interview_dir(), project_name)
    
    os.makedirs(project_dir, exist_ok=True)
    return project_dir


def get_transcript_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取转录文件目录"""
    project_dir = get_project_dir(project_name, evaluation_type)
    transcript_dir = os.path.join(project_dir, "transcriptions")
    os.makedirs(transcript_dir, exist_ok=True)
    return transcript_dir


def get_eval_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取评估结果目录"""
    project_dir = get_project_dir(project_name, evaluation_type)
    eval_dir = os.path.join(project_dir, "eval")
    os.makedirs(eval_dir, exist_ok=True)
    return eval_dir


def get_qa_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取问答对目录"""
    project_dir = get_project_dir(project_name, evaluation_type)
    qa_dir = os.path.join(project_dir, "qa_cache")
    os.makedirs(qa_dir, exist_ok=True)
    return qa_dir


def get_audio_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取音频文件目录"""
    project_dir = get_project_dir(project_name, evaluation_type)
    audio_dir = os.path.join(project_dir, "audio")
    os.makedirs(audio_dir, exist_ok=True)
    return audio_dir


def get_resume_dir(project_name: str, evaluation_type: str = None) -> str:
    """获取简历文件目录"""
    project_dir = get_project_dir(project_name, evaluation_type)
    # 支持多种可能的简历目录名称
    for resume_dir_name in ["简历", "resumes", "resume"]:
        resume_dir = os.path.join(project_dir, resume_dir_name)
        if os.path.exists(resume_dir):
            return resume_dir
    # 默认返回"简历"目录
    return os.path.join(project_dir, "简历")


def check_resume_exists(project_name: str, candidate_name: str, evaluation_type: str = None) -> bool:
    """检查候选人是否有简历文件"""
    try:
        resume_dir = get_resume_dir(project_name, evaluation_type)
        if not os.path.exists(resume_dir):
            return False
        
        # 支持的简历文件扩展名
        resume_extensions = ['.pdf', '.doc', '.docx', '.txt', '.md']
        
        for filename in os.listdir(resume_dir):
            # 检查文件名是否包含候选人姓名
            if candidate_name in filename:
                # 检查是否是支持的简历格式
                if any(filename.lower().endswith(ext) for ext in resume_extensions):
                    return True
        
        return False
    except Exception as e:
        logger.error(f"检查简历存在性失败: {e}")
        return False


# ============ 缓存管理 ============

def load_evaluation_cache(project: str, candidate_name: str, evaluation_type: str = None) -> Optional[dict]:
    """加载评估缓存"""
    try:
        eval_dir = get_eval_dir(project, evaluation_type)
        cache_file = os.path.join(eval_dir, f"{candidate_name}.json")
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"加载评估缓存失败: {e}")
    return None


def save_evaluation_cache(project: str, candidate_name: str, evaluation: dict, evaluation_type: str = None):
    """保存评估缓存"""
    try:
        eval_dir = get_eval_dir(project, evaluation_type)
        cache_file = os.path.join(eval_dir, f"{candidate_name}.json")
        cache_data = {
            "candidate_name": candidate_name,
            "project": project,
            "evaluation": evaluation,
            "cached_at": datetime.now().isoformat()
        }
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        logger.info(f"评估结果已缓存: {cache_file}")
    except Exception as e:
        logger.error(f"保存评估缓存失败: {e}")


def load_transcription_cache(project: str, filename: str) -> Optional[dict]:
    """加载转录缓存"""
    try:
        cache_dir = get_transcript_dir(project)
        cache_file = os.path.join(cache_dir, f"{filename}.json")
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.warning(f"读取转录缓存失败: {e}")
    return None


def save_transcription_cache(project: str, filename: str, candidate_name: str, transcript: str):
    """保存转录缓存"""
    try:
        cache_dir = get_transcript_dir(project)
        cache_file = os.path.join(cache_dir, f"{filename}.json")
        cache_data = {
            "file_name": filename,
            "candidate_name": candidate_name,
            "transcription": transcript,
            "processed_at": datetime.now().isoformat(),
            "cached": True
        }
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)
        logger.info(f"转录结果已缓存: {cache_file}")
    except Exception as e:
        logger.error(f"保存转录缓存失败: {e}")


# ============ 问答对提取 ============

def split_transcription(transcription: str) -> List[str]:
    """将转录文本分割成段落"""
    segments = []
    sentences = re.split(r'[。！？\n]+', transcription)
    current_segment = ""
    for sent in sentences:
        sent = sent.strip()
        if len(sent) < 5:
            continue
        if len(current_segment) < 200:
            current_segment += sent + "。"
        else:
            if current_segment:
                segments.append(current_segment)
            current_segment = sent + "。"
    if current_segment:
        segments.append(current_segment)
    return segments if segments else [transcription]


def extract_questions_from_text(text: str) -> List[str]:
    """从文本中提取问题列表"""
    questions = []
    question_keywords = [
        '请问', '想了解', '想知道', '什么是', '为什么', '如何', '怎么',
        '能否', '可以', '你觉得', '你认为', '谈谈', '介绍一下', '说一下',
        '讲讲', '分享', '描述', '举例', '解释', '说明', '哪个', '哪些',
        '何时', '何地', '何人', '何种', '哪方面', '怎么样', '什么原因',
        '如何看待', '怎么理解', '有什么', '做过什么', '负责什么',
        '职业规划', '未来打算', '期望薪资', '薪资要求'
    ]
    
    sentences = re.split(r'[。！？]', text)
    sentences = [s.strip() for s in sentences if s.strip() and len(s.strip()) > 5]
    
    for sentence in sentences:
        if any(keyword in sentence for keyword in question_keywords):
            questions.append(sentence)
    
    return list(set(questions))[:15]


def auto_extract_qa_pairs(transcript: str, project_dir: str = None) -> List[dict]:
    """自动从转录文本中提取问答对"""
    qa_pairs = []
    
    if not transcript or len(transcript) < 50:
        return [{
            "question": "面试交流",
            "answer": transcript[:1000] if transcript else "暂无内容",
            "category": "通用",
            "evaluation_points": "综合考察候选人的表达能力和经验",
            "start_time": 0,
            "end_time": 0
        }]
    
    questions_from_file = []
    if project_dir and os.path.exists(project_dir):
        questions_from_file = extract_questions_from_file(project_dir)
    
    if questions_from_file:
        for i, question in enumerate(questions_from_file):
            next_question = questions_from_file[i + 1] if i + 1 < len(questions_from_file) else None
            answer = extract_answer_from_transcript(transcript, question, next_question)
            
            if answer and len(answer) > 20:
                qa_pairs.append({
                    "question": question,
                    "answer": answer,
                    "category": "结构化问题",
                    "evaluation_points": "",
                    "start_time": 0,
                    "end_time": 0
                })
    
    if not qa_pairs:
        question_keywords = [
            '请问', '想了解', '想知道', '什么是', '为什么', '如何', '怎么',
            '能否', '可以', '你觉得', '你认为', '谈谈', '介绍一下', '说一下',
            '讲讲', '分享', '描述', '举例', '解释', '说明', '哪个', '哪些',
            '何时', '何地', '何人', '何种', '哪方面', '怎么样', '什么原因',
            '如何看待', '怎么理解', '有什么', '做过什么', '负责什么',
            '职业规划', '未来打算', '期望薪资', '薪资要求'
        ]
        
        sentences = re.split(r'[。！？]', transcript)
        sentences = [s.strip() for s in sentences if s.strip() and len(s.strip()) > 5]
        
        current_question = None
        current_answer = []
        
        for sentence in sentences:
            is_question = any(keyword in sentence for keyword in question_keywords)
            
            if is_question:
                if current_question and current_answer:
                    answer_text = '。'.join(current_answer)
                    if len(answer_text) > 20:
                        qa_pairs.append({
                            "question": current_question[:80] + "..." if len(current_question) > 80 else current_question,
                            "answer": answer_text[:500] + "..." if len(answer_text) > 500 else answer_text,
                            "category": "自动识别",
                            "evaluation_points": "",
                            "start_time": 0,
                            "end_time": 0
                        })
                
                current_question = sentence
                current_answer = []
            elif current_question:
                current_answer.append(sentence)
        
        if current_question and current_answer:
            answer_text = '。'.join(current_answer)
            if len(answer_text) > 20:
                qa_pairs.append({
                    "question": current_question[:80] + "..." if len(current_question) > 80 else current_question,
                    "answer": answer_text[:500] + "..." if len(answer_text) > 500 else answer_text,
                    "category": "自动识别",
                    "evaluation_points": "",
                    "start_time": 0,
                    "end_time": 0
                })
    
    if not qa_pairs:
        chunks = [transcript[i:i+500] for i in range(0, len(transcript), 500)]
        for i, chunk in enumerate(chunks[:5]):
            qa_pairs.append({
                "question": f"面试内容片段{i+1}",
                "answer": chunk,
                "category": "通用",
                "evaluation_points": "综合考察候选人的表达能力和经验",
                "start_time": 0,
                "end_time": 0
            })
    
    return qa_pairs[:15]


def extract_answer_from_transcript(transcript: str, question: str, next_question: str = None) -> str:
    """从转录文本中提取特定问题的答案并进行提炼"""
    if not transcript or not question:
        return ""
    
    start_idx = transcript.find(question)
    if start_idx == -1:
        for q_part in question[:-10].split('，')[:2]:
            if len(q_part) > 5:
                idx = transcript.find(q_part)
                if idx != -1:
                    start_idx = idx
                    break
    
    if start_idx == -1:
        return ""
    
    start_idx += len(question)
    
    if next_question and next_question in transcript:
        end_idx = transcript.find(next_question, start_idx)
    else:
        end_idx = len(transcript)
    
    raw_answer = transcript[start_idx:end_idx].strip()
    
    if not raw_answer:
        return ""
    
    raw_answer = re.sub(r'[。！？]+', '。', raw_answer)
    answer_parts = raw_answer.split('。')
    answer_parts = [p.strip() for p in answer_parts if p.strip()]
    
    if len(answer_parts) <= 3:
        refined_answer = '。'.join(answer_parts)
    else:
        refined_answer = '。'.join(answer_parts[:5])
    
    refined_answer = refined_answer[:500]
    if len(refined_answer) > 480:
        refined_answer = refined_answer[:480] + "..."
    
    return refined_answer.strip()


def extract_questions_from_file(project_dir: str) -> List[str]:
    """从项目目录中提取面试问题列表"""
    questions = []
    
    try:
        question_files = []
        if os.path.exists(project_dir):
            for f in os.listdir(project_dir):
                if f.lower().endswith(('.docx', '.doc', '.txt', '.pdf', '.xlsx', '.xls')):
                    if '面试题' in f or '问题' in f or 'question' in f.lower() or '结构化' in f:
                        question_files.append(os.path.join(project_dir, f))
        
        for q_file in question_files[:3]:
            try:
                with open(q_file, 'rb') as f:
                    content = f.read()
                
                text = ""
                if q_file.lower().endswith('.txt'):
                    text = content.decode('utf-8', errors='ignore')
                elif q_file.lower().endswith('.docx'):
                    docx_text = ""
                    try:
                        from docx import Document
                        import io
                        doc = Document(io.BytesIO(content))
                        docx_text = '\n'.join([para.text for para in doc.paragraphs])
                    except Exception:
                        pass
                    text = docx_text
                elif q_file.lower().endswith('.xlsx') or q_file.lower().endswith('.xls'):
                    excel_text = ""
                    try:
                        import io
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(content))
                        for sheet in wb.sheetnames:
                            ws = wb[sheet]
                            for row in ws.iter_rows(values_only=True):
                                for cell in row:
                                    if cell and isinstance(cell, str):
                                        excel_text += cell + '\n'
                    except Exception:
                        try:
                            import xlrd
                            wb = xlrd.open_workbook(file_contents=content)
                            for sheet in wb.sheets():
                                for row in sheet.get_rows():
                                    for cell in row:
                                        if cell.value:
                                            excel_text += str(cell.value) + '\n'
                        except Exception:
                            pass
                    text = excel_text
                
                if text:
                    lines = text.split('\n')
                    for line in lines:
                        line = line.strip()
                        if line and len(line) > 10:
                            if line[0].isdigit() and ('.' in line[:5] or '、' in line[:5]):
                                parts = re.split(r'[.。、]\s*', line, maxsplit=1)
                                if len(parts) > 1:
                                    question = parts[1].strip()
                                    if len(question) > 10:
                                        questions.append(question)
                            elif any(keyword in line for keyword in ['请', '谈谈', '介绍', '分享', '说明', '分析', '如何', '为什么', '什么']):
                                questions.append(line)
            except Exception:
                continue
        
        questions = list(set(questions))[:15]
    
    except Exception as e:
        logger.warning(f"提取面试问题失败: {e}")
    
    return questions


async def save_qa_result(project: str, candidate_name: str, qa_pairs: List[dict], questions: List[dict], evaluation_type: str = None):
    """保存问答对结果"""
    try:
        qa_dir = get_qa_dir(project, evaluation_type)
        qa_file = os.path.join(qa_dir, f"{candidate_name}.json")
        result = {
            "qa_pairs": qa_pairs,
            "questions": questions,
            "timestamp": datetime.now().isoformat()
        }
        with open(qa_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        logger.info(f"问答对已保存: {qa_file}")
    except Exception as e:
        logger.error(f"保存问答对失败: {e}")


# ============ 候选人姓名提取 ============

def extract_candidate_name(filename: str) -> str:
    """从文件名提取候选人姓名"""
    try:
        name_part = filename.split('_')[-1]
        name = name_part.split('.')[0]
        for i, char in enumerate(name):
            if not char.isdigit():
                return name[i:]
        return name
    except:
        return filename


# ============ 项目管理 ============

def get_projects(evaluation_type: str = None) -> List[str]:
    """获取项目列表"""
    projects = []
    
    try:
        if evaluation_type:
            type_dir = get_type_dir(evaluation_type)
            if os.path.exists(type_dir):
                for item in os.listdir(type_dir):
                    item_path = os.path.join(type_dir, item)
                    if os.path.isdir(item_path) and not item.startswith('.') and item != 'eval':
                        projects.append(item)
        else:
            for type_name in ['员工', '干部']:
                type_dir = os.path.join(get_base_interview_dir(), type_name)
                if os.path.exists(type_dir):
                    for item in os.listdir(type_dir):
                        item_path = os.path.join(type_dir, item)
                        if os.path.isdir(item_path) and not item.startswith('.') and item != 'eval':
                            projects.append(f"{type_name}/{item}")
    except Exception as e:
        logger.error(f"获取项目列表失败: {e}")
    
    return sorted(projects)